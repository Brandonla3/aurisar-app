#!/usr/bin/env python3
"""
Aurisar Exercise Catalog Audit
==============================
Builds a full Excel audit workbook of the runtime exercise catalog.

The runtime catalog is the union of:
  - the app bundle  (src/data/exercises.js, exported to JSON via Node)
  - the database    (Supabase `exercises` table, exported to JSON)
merged exactly the way src/utils/exerciseLibrary.js:loadExercises() does it:
  - IDs present in both: the database overrides equipment, difficulty,
    pbType, pbTier, xpClassMap; everything else keeps the bundled value.
  - IDs only in the database are appended (the app does NOT filter on the
    database `archived` flag, so archived entries are part of the runtime
    catalog too).

Inputs (produced beforehand):
  /tmp/audit/static_exercises.json   - bundled EXERCISES array
  /tmp/audit/supabase_exercises.json - full DB table
  /tmp/audit/classes.json            - CLASSES definition

Output:
  Aurisar_Exercise_Audit_<date>.xlsx
"""

import json
import re
from collections import Counter, defaultdict
from datetime import date
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STATIC_PATH = "/tmp/audit/static_exercises.json"
DB_PATH = "/tmp/audit/supabase_exercises.json"
OUT_PATH = f"Aurisar_Exercise_Audit_{date.today().isoformat()}.xlsx"

CLASS_KEYS = ["warrior", "gladiator", "warden", "phantom", "tempest",
              "warlord", "druid", "oracle", "titan", "striker", "alchemist"]
KNOWN_CATEGORIES = {"strength", "cardio", "flexibility", "endurance"}
KNOWN_DIFFICULTIES = {"Beginner", "Intermediate", "Advanced"}
KNOWN_MUSCLE_GROUPS = {"abs", "back", "bicep", "calves", "cardio", "chest",
                       "forearm", "full_body", "glutes", "legs", "shoulder", "tricep"}
WEIGHTED_EQUIPMENT = {"barbell", "dumbbell", "kettlebell", "machine", "cable",
                      "e-z curl bar", "ez curl bar", "medicine ball", "landmine",
                      "trap bar", "smith machine", "weight plate"}
# exerciseType is a comma-separated tag list; the scoring category is
# consistent if at least one tag maps to it
CATEGORY_TAGS = {
    "strength": {"strength", "calisthenics", "core", "isometric", "functional",
                 "rehabilitation", "plyometric"},
    "flexibility": {"flexibility", "stretching", "yoga", "mobility", "balance",
                    "cooldown", "warmup"},
    "cardio": {"cardio", "hiit", "plyometric", "warmup", "endurance"},
    "endurance": {"endurance", "cardio"},
}

# ---------------------------------------------------------------- load + merge

def map_db_row(ex):
    """Maps a database row to the app exercise shape. The id/name/category/
    equipment/difficulty/PB/xpClassMap mappings replicate
    exerciseLibrary.js:loadExercises(); the remaining fields are carried for
    audit completeness (the table holds them even though the app's mapper
    drops them for database-only entries)."""
    mg = ex.get("muscle_group") or "back"
    return {
        "id": ex.get("id"),
        "name": ex.get("name"),
        "category": ex.get("category") or "strength",
        "secondaryCategory": None,
        "muscleGroup": mg,
        "icon": ex.get("icon") or "🏋️",
        "baseXP": ex.get("base_xp") or 40,
        "muscles": mg[:1].upper() + mg[1:] if ex.get("muscle_group") else "",
        "desc": ex.get("description") or "",
        "equipment": ex.get("equipment") or "bodyweight",
        "difficulty": ex.get("difficulty") or "Intermediate",
        "classAffinity": ex.get("class_affinity") or "all",
        "exerciseType": ex.get("exercise_type") or "",
        "pbType": ex.get("pb_type") or None,
        "pbTier": ex.get("pb_tier") or "Personal",
        "primaryPBMetric": ex.get("primary_pb_metric"),
        "secondaryPB": ex.get("secondary_pb"),
        "markAsPB": ex.get("mark_as_pb"),
        "wodViable": bool(ex.get("wod_viable")),
        "compound": bool(ex.get("compound")),
        "calisthenics": bool(ex.get("calisthenics")),
        "olympic": bool(ex.get("olympic")),
        "plyometric": bool(ex.get("plyometric")),
        "isolation": bool(ex.get("isolation")),
        "tracksWeight": bool(ex.get("tracks_weight")),
        "tracksDistance": bool(ex.get("tracks_distance")),
        "tracksInclineSpeed": bool(ex.get("tracks_incline_speed") or ex.get("tracks_incline")),
        "resistanceLevel": bool(ex.get("resistance_level")),
        "xpInputFormula": ex.get("xp_input_formula"),
        "supersetEligible": ex.get("superset_eligible"),
        "intervalEligible": ex.get("interval_eligible"),
        "defaultSets": ex.get("default_sets"),
        "defaultReps": ex.get("default_reps"),
        "defaultDurationMin": ex.get("default_duration_min"),
        "archived": bool(ex.get("archived")),
        "xpClassMap": ex.get("xp_class_map") or {},
    }


# fields the app NEVER takes from the DB for bundled IDs -> a divergence means
# the catalog DB value is silently ignored at runtime
COMPARE_BUNDLE_WINS = [
    "name", "category", "muscleGroup", "icon", "baseXP", "desc",
    "classAffinity", "exerciseType", "xpInputFormula", "primaryPBMetric",
    "markAsPB", "wodViable", "compound", "isolation", "calisthenics",
    "olympic", "plyometric", "tracksWeight", "tracksDistance",
    "tracksInclineSpeed", "resistanceLevel", "supersetEligible",
    "intervalEligible",
]
# fields the app DOES override from the DB
COMPARE_DB_WINS = ["equipment", "difficulty", "pbType", "pbTier", "xpClassMap"]


def _norm_cmp(v):
    if isinstance(v, str):
        return " ".join(v.split())
    return v


def load_and_merge():
    static = json.load(open(STATIC_PATH))
    db_raw = json.load(open(DB_PATH))
    db = {r["id"]: r for r in db_raw if r.get("id")}

    merged = []
    divergences = []  # (id, field, bundle_value, db_value, winner)
    static_ids = set()

    for ex in static:
        static_ids.add(ex["id"])
        row = dict(ex)
        row["inBundle"] = True
        row["inDatabase"] = ex["id"] in db
        row["archived"] = False
        if ex["id"] in db:
            d = map_db_row(db[ex["id"]])
            row["archived"] = d["archived"]
            # secondaryPB and the logging defaults live only in the catalog
            # database; loadExercises() does not patch them onto bundled
            # exercises. They are included for audit completeness and the
            # Data Dictionary notes that the app drops them at runtime.
            row["secondaryPB"] = d["secondaryPB"]
            for f in COMPARE_BUNDLE_WINS:
                bv, dv = ex.get(f), d.get(f)
                if f not in ex:
                    continue  # legacy-shape records are flagged separately
                if dv in (None, ""):
                    continue
                if isinstance(bv, bool) or isinstance(dv, bool):
                    if bool(bv) != bool(dv):
                        divergences.append((ex["id"], f, bv, dv, "bundle"))
                elif _norm_cmp(bv) != _norm_cmp(dv):
                    divergences.append((ex["id"], f, bv, dv, "bundle"))
            for f in COMPARE_DB_WINS:
                bv, dv = ex.get(f), d.get(f)
                if f in ex and bv is not None and bv != dv:
                    divergences.append((ex["id"], f, bv, dv, "database"))
                row[f] = dv
            for f in ["defaultSets", "defaultReps", "defaultDurationMin"]:
                row[f] = d[f]
        merged.append(row)

    for ex_id, raw in db.items():
        if ex_id in static_ids:
            continue
        row = map_db_row(raw)
        row["inBundle"] = False
        row["inDatabase"] = True
        merged.append(row)

    return merged, divergences, static, db_raw


# ------------------------------------------------------------ duplicate groups

SYNONYMS = {
    "pushups": "pushup", "push-up": "pushup", "pressup": "pushup",
    "situps": "situp", "sit-up": "situp",
    "pullups": "pullup", "pull-up": "pullup",
    "chinups": "chinup", "chin-up": "chinup",
    "stepups": "stepup", "step-up": "stepup",
    "db": "dumbbell", "dumbell": "dumbbell", "dumbbells": "dumbbell",
    "bb": "barbell", "barbells": "barbell",
    "kb": "kettlebell", "kettlebells": "kettlebell",
    "alt": "alternating", "alternate": "alternating",
    "extention": "extension", "extentions": "extension",
    "oly": "olympic",
    "childs": "child", "child's": "child",
}
NO_DEPLURAL = {"abs", "biceps", "triceps", "quads", "glutes", "lats", "delts",
               "calves", "obliques", "hamstrings", "press", "cross", "pass"}
# qualifier tokens that make two near-identical names DIFFERENT exercises
QUALIFIERS = {"incline", "decline", "flat", "close", "wide", "narrow", "reverse",
              "front", "back", "rear", "single", "one", "two", "double", "half",
              "full", "seated", "standing", "lying", "kneeling", "bent", "straight",
              "alternating", "left", "right", "high", "low", "inner", "outer",
              "overhead", "underhand", "overhand", "sumo", "deficit", "paused",
              "weighted", "assisted", "negative", "side", "behind", "smith",
              "swiss", "stability", "barbell", "dumbbell", "kettlebell", "cable",
              "machine", "band", "banded", "bodyweight", "plate", "landmine",
              "ring", "trx", "elevated", "jump", "jumping", "explosive", "iso",
              "isometric", "partial", "tempo", "speed", "walking", "static",
              "arm", "leg", "grip", "stance", "up", "down", "in", "out", "to"}


def depluralize(tok):
    if tok in NO_DEPLURAL or len(tok) <= 3:
        return tok
    if tok.endswith("ies"):
        return tok[:-3] + "y"
    if tok.endswith("s") and not tok.endswith("ss"):
        return tok[:-1]
    return tok


def norm_tokens(name):
    s = (name or "").lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[/\-_,+()]", " ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    toks = []
    for t in s.split():
        t = SYNONYMS.get(t, t)
        t = depluralize(t)
        t = SYNONYMS.get(t, t)
        if t:
            toks.append(t)
    return toks


def norm_key(name):
    return " ".join(sorted(norm_tokens(name)))


def directional_sig(tokens):
    """('high', 'low') for 'high to low wood chop' — used to avoid merging
    'High to Low X' with 'Low to High X' (different movement directions)."""
    if "to" in tokens:
        i = tokens.index("to")
        if 0 < i < len(tokens) - 1:
            return (tokens[i - 1], tokens[i + 1])
    return None


SIDE_TOKENS = {"l", "r", "left", "right"}
# token pairs that look like spelling variants but are opposite movements
ANTONYM_PAIRS = {frozenset(p) for p in [
    ("abductor", "adductor"), ("abduction", "adduction"),
    ("supination", "pronation"), ("internal", "external"),
    ("flexion", "extension"), ("concentric", "eccentric"),
]}
ROMAN_OR_NUM = re.compile(r"^(?:[ivx]+|\d+)$")


class UnionFind:
    def __init__(self):
        self.parent = {}

    def find(self, x):
        self.parent.setdefault(x, x)
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[rb] = ra


def fuzzy_same(a_toks, b_toks):
    """True only for high-confidence matches: identical token sets, or
    differences that are pure spelling variants of one another."""
    sa, sb = set(a_toks), set(b_toks)
    if sa == sb:
        return True
    da, db_ = sorted(sa - sb), sorted(sb - sa)
    if len(da) != len(db_) or not da or len(da) > 2:
        return False
    for x, y in zip(da, db_):
        if x in QUALIFIERS or y in QUALIFIERS:
            return False
        if frozenset((x, y)) in ANTONYM_PAIRS:
            return False
        if ROMAN_OR_NUM.match(x) or ROMAN_OR_NUM.match(y):
            return False
        if SequenceMatcher(None, x, y).ratio() < 0.8 or abs(len(x) - len(y)) > 2:
            return False
        if sum(1 for c1, c2 in zip(x, y) if c1 != c2) + abs(len(x) - len(y)) > 2:
            return False
    return True


def find_duplicates(merged):
    uf = UnionFind()
    basis = {}  # group-member id -> match basis notes

    # pass 1: exact normalized-name key (blocking 'A to B' vs 'B to A')
    by_key = defaultdict(list)
    toks_by_id = {}
    for ex in merged:
        by_key[norm_key(ex["name"])].append(ex["id"])
        toks_by_id[ex["id"]] = norm_tokens(ex["name"])
    for ids in by_key.values():
        for i in range(len(ids)):
            for j in range(i + 1, len(ids)):
                sa = directional_sig(toks_by_id[ids[i]])
                sb = directional_sig(toks_by_id[ids[j]])
                if sa and sb and sa != sb and sa == (sb[1], sb[0]):
                    continue  # e.g. High-to-Low vs Low-to-High wood chop
                uf.union(ids[i], ids[j])

    # pass 2: parenthetical handling (requires same muscle group)
    #   a) inner alias: 'Bodyweight Squat (air squat)' <-> 'Air Squat'
    #   b) side marker: 'Split squat (L)' <-> 'Split Squats'
    #   c) base-name variant: 'Machine Shoulder (Military) Press' <-> 'Machine
    #      Shoulder Press' — skipped when the parenthetical carries a movement
    #      qualifier (e.g. '(Clean Grip)', '(Double Unders)', '(No jump)')
    mg = {e["id"]: e.get("muscleGroup") for e in merged}
    for ex in merged:
        name = ex.get("name") or ""
        inners = re.findall(r"\(([^)]+)\)", name)
        if not inners:
            continue
        for inner in inners:
            k = norm_key(inner)
            if len(k) >= 4:
                for other_id in by_key.get(k, []):
                    if other_id != ex["id"] and mg.get(other_id) == mg.get(ex["id"]):
                        uf.union(ex["id"], other_id)
                        basis.setdefault(ex["id"], "alias")
                        basis.setdefault(other_id, "alias")
        base = re.sub(r"\([^)]*\)", " ", name).strip()
        bkey = norm_key(base)
        if not base or not bkey:
            continue
        inner_toks = {t for innr in inners for t in norm_tokens(innr)}
        if inner_toks and inner_toks <= SIDE_TOKENS:
            kind = "side"
        elif inner_toks and all(ROMAN_OR_NUM.match(t) for t in inner_toks):
            kind = "variant"  # junk numeric suffix like '(1)'
        elif not inner_toks or (inner_toks & QUALIFIERS):
            continue
        else:
            kind = "variant"
        for other_id in by_key.get(bkey, []):
            if other_id != ex["id"] and mg.get(other_id) == mg.get(ex["id"]):
                uf.union(ex["id"], other_id)
                basis[ex["id"]] = kind
                basis[other_id] = basis.get(other_id) or kind

    # pass 3: spelling-variant fuzzy match within same muscleGroup+equipment
    buckets = defaultdict(list)
    for ex in merged:
        buckets[(ex.get("muscleGroup"), ex.get("equipment"))].append(ex)
    for bucket in buckets.values():
        toks = [(e["id"], norm_tokens(e["name"])) for e in bucket]
        for i in range(len(toks)):
            for j in range(i + 1, len(toks)):
                if uf.find(toks[i][0]) == uf.find(toks[j][0]):
                    continue
                sa, sb = directional_sig(toks[i][1]), directional_sig(toks[j][1])
                if sa and sb and sa != sb and sa == (sb[1], sb[0]):
                    continue
                if fuzzy_same(toks[i][1], toks[j][1]):
                    uf.union(toks[i][0], toks[j][0])
                    basis[toks[i][0]] = "fuzzy"
                    basis[toks[j][0]] = "fuzzy"

    groups = defaultdict(list)
    for ex in merged:
        groups[uf.find(ex["id"])].append(ex["id"])
    groups = {k: v for k, v in groups.items() if len(v) > 1}
    return groups, basis


def completeness_score(ex):
    score = 0
    if not ex.get("archived"):
        score += 200
    if ex.get("inBundle") and ex.get("inDatabase"):
        score += 50
    if ex.get("inBundle"):
        score += 20
    if ex.get("xpInputFormula"):
        score += 10
    score += min(len(ex.get("desc") or ""), 400) / 400 * 10
    score += sum(1 for f in ("equipment", "difficulty", "pbType", "exerciseType")
                 if ex.get(f))
    return score


# ------------------------------------------------------------- inconsistencies

def check_inconsistencies(merged, divergences, static_raw, db_raw):
    findings = []  # (severity, ex_id, name, issue, detail)

    def add(sev, ex, issue, detail):
        findings.append((sev, ex["id"], ex.get("name") or "", issue, detail))

    # duplicate IDs inside each layer
    for label, rows in (("app bundle", static_raw), ("database", db_raw)):
        c = Counter(r["id"] for r in rows if r.get("id"))
        for ex_id, n in c.items():
            if n > 1:
                findings.append(("High", ex_id, "", "Duplicate ID",
                                 f"ID appears {n} times in the {label}"))

    xp_by_cat = defaultdict(list)
    for ex in merged:
        xp_by_cat[ex.get("category")].append(ex.get("baseXP") or 0)
    xp_med = {c: sorted(v)[len(v) // 2] for c, v in xp_by_cat.items()}

    for ex in merged:
        cat = ex.get("category")
        etype = ex.get("exerciseType")
        mg = ex.get("muscleGroup")
        diff = ex.get("difficulty")
        eq = (ex.get("equipment") or "").lower()
        xmap = ex.get("xpClassMap") or {}
        formula = ex.get("xpInputFormula")
        name = ex.get("name") or ""

        if ex.get("archived"):
            if ex.get("inBundle"):
                add("High", ex, "Archived but still live in app",
                    "Marked archived in the catalog database but also ships in the "
                    "app bundle, so it remains fully active")
            else:
                add("High", ex, "Archived but still live in app",
                    "Marked archived in the catalog database, but the app's loader "
                    "does not filter archived entries, so it still appears in the app")

        if "exerciseType" not in ex and "wodViable" not in ex:
            add("High", ex, "Missing extended metadata",
                "No exerciseType/equipment/difficulty/flags/PB/XP-formula fields "
                "(legacy record shape)")
        if not (ex.get("desc") or "").strip():
            add("Medium", ex, "Missing description", "Description is empty")
        if not (ex.get("icon") or "").strip():
            add("Medium", ex, "Missing icon", "Icon is empty")
        if not name.strip():
            add("High", ex, "Missing name", "Name is empty")

        if re.search(r"_|\s{2,}|\s-\w|[a-z]\($|-\s*\d+\s*$", name):
            add("Low", ex, "Name formatting issue",
                f"Display name looks malformed: {name!r}")

        d = (ex.get("desc") or "").strip()
        if d and d[-1] not in ".!?)\"'":
            add("Low", ex, "Description may be truncated",
                f"Ends without terminal punctuation: …{d[-60:]!r}")

        if cat not in KNOWN_CATEGORIES:
            add("Medium", ex, "Unknown category", f"category={cat!r}")
        if mg not in KNOWN_MUSCLE_GROUPS:
            add("Medium", ex, "Unknown muscle group", f"muscleGroup={mg!r}")
        if diff is not None and diff not in KNOWN_DIFFICULTIES:
            add("Medium", ex, "Unknown difficulty", f"difficulty={diff!r}")
        aff = ex.get("classAffinity")
        if aff is not None and aff not in CLASS_KEYS and aff != "all":
            add("Medium", ex, "Unknown class affinity", f"classAffinity={aff!r}")

        if etype and cat in CATEGORY_TAGS:
            tags = {t.strip().lower() for t in etype.split(",") if t.strip()}
            if tags and not (tags & CATEGORY_TAGS[cat]):
                add("Medium", ex, "Category not reflected in type tags",
                    f"category={cat!r} but exerciseType tags are {etype!r}")

        if xmap:
            missing = [k for k in CLASS_KEYS if k not in xmap]
            if missing:
                add("Medium", ex, "Incomplete XP class map",
                    f"Missing multipliers for: {', '.join(missing)}")
            if aff and aff in xmap and aff != "all":
                top = max(xmap.values())
                if xmap[aff] < top - 1e-9:
                    best = [k for k, v in xmap.items() if v >= top - 1e-9]
                    add("Low", ex, "Affinity is not top XP class",
                        f"classAffinity={aff} ({int(round(xmap[aff]*100))}%) but "
                        f"highest XP class is {'/'.join(best)} ({int(round(top*100))}%)")
        elif "xpClassMap" in ex or ex.get("inDatabase"):
            add("Medium", ex, "Missing XP class map",
                "No per-class XP multipliers; XP falls back to category bonus")

        if ex.get("compound") and ex.get("isolation"):
            add("Medium", ex, "Compound AND isolation", "Both flags are true")

        if "tracksWeight" in ex:
            if ex.get("tracksWeight") and eq == "bodyweight":
                add("Medium", ex, "Weight tracking on bodyweight exercise",
                    "tracksWeight=true but equipment=bodyweight")
            if not ex.get("tracksWeight") and eq in WEIGHTED_EQUIPMENT:
                add("Medium", ex, "No weight tracking on weighted exercise",
                    f"tracksWeight=false but equipment={eq}")

        if formula:
            if "weight" in formula.lower() and not ex.get("tracksWeight"):
                add("Medium", ex, "XP formula/tracking mismatch",
                    f"Formula {formula!r} uses weight but tracksWeight=false")
            if "distance" in formula.lower() and not ex.get("tracksDistance"):
                add("Medium", ex, "XP formula/tracking mismatch",
                    f"Formula {formula!r} uses distance but tracksDistance=false")

        pb = (ex.get("pbType") or "").lower()
        if "weight" in pb and not ex.get("tracksWeight"):
            add("Low", ex, "PB type/tracking mismatch",
                f"pbType={ex.get('pbType')!r} but tracksWeight=false")

        spb = ex.get("secondaryPB")
        if spb and "miles)" in spb and "or km" not in spb:
            add("Low", ex, "Inconsistent secondary PB label",
                f"secondaryPB={spb!r}; most exercises use "
                "'Best Total Distance (miles or km)'")

        bxp = ex.get("baseXP")
        if isinstance(bxp, (int, float)) and cat in xp_med:
            if bxp < 10 or bxp > 60 or abs(bxp - xp_med[cat]) > 20:
                add("Low", ex, "Base XP outlier",
                    f"baseXP={bxp} (median for {cat}: {xp_med[cat]})")

    by_ex = defaultdict(list)
    for ex_id, field, bv, dv, wins in divergences:
        by_ex[ex_id].append((field, bv, dv, wins))
    name_by_id = {e["id"]: e.get("name") or "" for e in merged}
    for ex_id, items in by_ex.items():
        ignored = []
        for f, bv, dv, w in items:
            if w != "bundle":
                continue
            if f == "desc" and isinstance(bv, str) and isinstance(dv, str) \
                    and len(bv) < len(dv):
                findings.append((
                    "Medium", ex_id, name_by_id.get(ex_id, ""),
                    "Description truncated in app",
                    f"App shows {len(bv)} chars but the catalog database holds the "
                    f"full {len(dv)}-char text (app cuts off at: …{bv[-60:]!r})"))
            else:
                ignored.append((f, bv, dv))
        if ignored:
            det = "; ".join(f"{f}: app shows {bv!r}, catalog record says {dv!r}"
                            for f, bv, dv in ignored)
            findings.append(("Medium", ex_id, name_by_id.get(ex_id, ""),
                             "Conflicting values between catalog layers (app keeps bundled value)",
                             det[:900]))
    return findings


# --------------------------------------------------------------- excel helpers

HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BAND_FILLS = [PatternFill("solid", fgColor="FFF2CC"),
              PatternFill("solid", fgColor="DDEBF7")]
SEV_FILLS = {"High": PatternFill("solid", fgColor="F8CBAD"),
             "Medium": PatternFill("solid", fgColor="FFE699"),
             "Low": PatternFill("solid", fgColor="E2EFDA")}


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------------------------- main

def main():
    merged, divergences, static_raw, db_raw = load_and_merge()
    n_total = len(merged)
    groups, basis = find_duplicates(merged)
    findings = check_inconsistencies(merged, divergences, static_raw, db_raw)

    by_id = {e["id"]: e for e in merged}
    group_of = {}
    group_num = {}
    canonical = {}
    ordered = sorted(groups.items(),
                     key=lambda kv: min((by_id[i].get("name") or "").lower() for i in kv[1]))
    for n, (root, ids) in enumerate(ordered, 1):
        gid = f"DUP-{n:03d}"
        group_num[root] = gid
        best = max(ids, key=lambda i: completeness_score(by_id[i]))
        for i in ids:
            group_of[i] = gid
        canonical[gid] = best

    issues_by_id = defaultdict(list)
    for sev, ex_id, _name, issue, _detail in findings:
        if ex_id:
            issues_by_id[ex_id].append(issue)

    # sort alphabetically, but anchor duplicate-group members together
    def sort_key(ex):
        gid = group_of.get(ex["id"])
        if gid:
            anchor = (by_id[canonical[gid]].get("name") or "").lower()
            return (anchor, 0 if ex["id"] == canonical[gid] else 1,
                    (ex.get("name") or "").lower())
        return ((ex.get("name") or "").lower(), 0, "")

    rows = sorted(merged, key=sort_key)

    wb = Workbook()

    # ---------------- Master List ----------------
    ws = wb.active
    ws.title = "Master List"
    headers = (["Dup Group", "Suggested Canonical", "ID", "Name", "Icon",
                "Category", "Secondary Category", "Exercise Type", "Muscle Group",
                "Muscles (display)", "Equipment", "Difficulty", "Description",
                "Base XP", "Class Affinity", "Best XP Class(es)"]
               + [f"{k.title()} XP %" for k in CLASS_KEYS]
               + ["XP Input Formula", "PB Type", "PB Tier", "Primary PB Metric",
                  "Secondary PB", "Mark As PB", "WOD Viable", "Compound",
                  "Isolation", "Calisthenics", "Olympic", "Plyometric",
                  "Tracks Weight", "Tracks Distance", "Tracks Incline/Speed",
                  "Resistance Level", "Superset Eligible", "Interval Eligible",
                  "Default Sets", "Default Reps", "Default Duration (min)",
                  "Archived", "In App Bundle", "In Database", "Issues"])
    ws.append(headers)
    style_header(ws, len(headers))

    def b(v):
        if v is None:
            return ""
        return "Yes" if v else "No"

    band_idx = {}
    for ex in rows:
        gid = group_of.get(ex["id"], "")
        if gid and gid not in band_idx:
            band_idx[gid] = len(band_idx) % 2
        xmap = ex.get("xpClassMap") or {}
        top = max(xmap.values()) if xmap else None
        best = "/".join(k for k in CLASS_KEYS if xmap.get(k, 0) >= (top or 0) - 1e-9) if xmap else ""
        issues = sorted(set(issues_by_id.get(ex["id"], [])))
        row = ([gid,
                ("Yes" if gid and canonical.get(gid) == ex["id"] else ("No" if gid else "")),
                ex["id"], ex.get("name"), ex.get("icon"),
                ex.get("category"), ex.get("secondaryCategory") or "",
                ex.get("exerciseType") or "", ex.get("muscleGroup"),
                ex.get("muscles") or "", ex.get("equipment") or "",
                ex.get("difficulty") or "", ex.get("desc") or "",
                ex.get("baseXP"), ex.get("classAffinity") or "", best]
               + [xmap.get(k) if k in xmap else None for k in CLASS_KEYS]
               + [ex.get("xpInputFormula") or "", ex.get("pbType") or "",
                  ex.get("pbTier") or "", ex.get("primaryPBMetric") or "",
                  ex.get("secondaryPB") or "",
                  b(ex.get("markAsPB")), b(ex.get("wodViable")), b(ex.get("compound")),
                  b(ex.get("isolation")), b(ex.get("calisthenics")), b(ex.get("olympic")),
                  b(ex.get("plyometric")), b(ex.get("tracksWeight")),
                  b(ex.get("tracksDistance")), b(ex.get("tracksInclineSpeed")),
                  b(ex.get("resistanceLevel")), b(ex.get("supersetEligible")),
                  b(ex.get("intervalEligible")), ex.get("defaultSets"),
                  ex.get("defaultReps"), ex.get("defaultDurationMin"),
                  b(ex.get("archived")), b(ex.get("inBundle")), b(ex.get("inDatabase")),
                  "; ".join(issues)])
        ws.append(row)
        r = ws.max_row
        if gid:
            fill = BAND_FILLS[band_idx[gid]]
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = fill
        if issues:
            ws.cell(row=r, column=len(headers)).fill = SEV_FILLS["Medium"]
        for c in range(17, 17 + len(CLASS_KEYS)):
            ws.cell(row=r, column=c).number_format = "0%"
        ws.cell(row=r, column=13).alignment = Alignment(wrap_text=False)

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    set_widths(ws, [9, 9, 26, 32, 5, 11, 10, 11, 12, 12, 13, 12, 60, 8, 12, 16]
               + [9] * len(CLASS_KEYS)
               + [18, 20, 9, 26, 26, 8, 8, 9, 8, 11, 8, 10, 8, 8, 9, 10, 9, 8,
                  8, 8, 9, 9, 9, 9, 50])

    # ---------------- Duplicate Groups ----------------
    ws2 = wb.create_sheet("Duplicate Groups")
    h2 = ["Group", "Match Basis", "Confidence", "# Members", "Member IDs",
          "Member Names", "Suggested Canonical", "Differences Between Members"]
    ws2.append(h2)
    style_header(ws2, len(h2))
    diff_fields = ["category", "exerciseType", "muscleGroup", "equipment",
                   "difficulty", "baseXP", "classAffinity", "icon",
                   "pbType", "archived", "inBundle", "inDatabase"]
    for root, ids in ordered:
        gid = group_num[root]
        members = [by_id[i] for i in ids]
        bases = {basis.get(i) for i in ids if basis.get(i)}
        if "side" in bases:
            basis_label, conf = "Left/Right side variants of same movement", \
                "Review — same movement, side-specific entries"
        elif "variant" in bases:
            basis_label, conf = "Base name + parenthetical variant", \
                "Review — likely same movement"
        elif "fuzzy" in bases:
            basis_label, conf = "Spelling/format variant", "High"
        elif "alias" in bases:
            basis_label, conf = "Name alias (parenthetical)", "High"
        else:
            basis_label, conf = "Same name (normalized)", "High"
        diffs = []
        for f in diff_fields:
            vals = {repr(m.get(f)) for m in members}
            if len(vals) > 1:
                diffs.append(f"{f}: " + " vs ".join(sorted(vals)))
        ws2.append([gid, basis_label, conf, len(ids),
                    ", ".join(sorted(ids)),
                    " | ".join(m.get("name") or "" for m in members),
                    canonical[gid],
                    "; ".join(diffs) if diffs else "Metadata identical"])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{ws2.max_row}"
    set_widths(ws2, [9, 30, 34, 10, 50, 60, 26, 80])

    # ---------------- Inconsistencies ----------------
    ws3 = wb.create_sheet("Inconsistencies")
    h3 = ["Severity", "Exercise ID", "Exercise Name", "Issue", "Detail"]
    ws3.append(h3)
    style_header(ws3, len(h3))
    sev_rank = {"High": 0, "Medium": 1, "Low": 2}
    for sev, ex_id, name, issue, detail in sorted(
            findings, key=lambda f: (sev_rank.get(f[0], 3), f[3], f[1])):
        ws3.append([sev, ex_id, name, issue, detail])
        ws3.cell(row=ws3.max_row, column=1).fill = SEV_FILLS.get(sev, SEV_FILLS["Low"])
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:E{ws3.max_row}"
    set_widths(ws3, [10, 28, 32, 38, 100])

    # ---------------- Summary ----------------
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Aurisar Exercise Catalog Audit — " + date.today().isoformat()])
    ws4["A1"].font = Font(bold=True, size=14)
    ws4.append([])

    def section(title, pairs):
        ws4.append([title])
        ws4.cell(row=ws4.max_row, column=1).font = Font(bold=True, size=11)
        for k, v in pairs:
            ws4.append(["", k, v])
        ws4.append([])

    n_bundle = sum(1 for e in merged if e["inBundle"])
    n_db = sum(1 for e in merged if e["inDatabase"])
    n_both = sum(1 for e in merged if e["inBundle"] and e["inDatabase"])
    n_arch = sum(1 for e in merged if e.get("archived"))
    dup_members = sum(len(v) for v in groups.values())
    section("Catalog totals", [
        ("Total exercises (runtime catalog)", n_total),
        ("In app bundle", n_bundle),
        ("In catalog database", n_db),
        ("In both layers", n_both),
        ("Bundle only", n_bundle - n_both),
        ("Database only", n_db - n_both),
        ("Marked archived (but still live in app)", n_arch),
    ])
    section("Duplicates", [
        ("Duplicate groups", len(groups)),
        ("Exercises involved in a duplicate group", dup_members),
        ("Potential redundant entries (members minus one per group)",
         dup_members - len(groups)),
    ])
    sev_counts = Counter(f[0] for f in findings)
    issue_counts = Counter(f[3] for f in findings)
    section("Inconsistencies", [("Total findings", len(findings))]
            + [(f"{s} severity", sev_counts.get(s, 0)) for s in ("High", "Medium", "Low")])
    section("Findings by issue type",
            sorted(issue_counts.items(), key=lambda kv: -kv[1]))

    for title, field in [("By category", "category"), ("By muscle group", "muscleGroup"),
                         ("By difficulty", "difficulty"), ("By equipment", "equipment"),
                         ("By class affinity", "classAffinity")]:
        c = Counter((e.get(field) or "(blank)") for e in merged)
        section(title, sorted(c.items(), key=lambda kv: -kv[1]))
    c = Counter(e.get("baseXP") for e in merged)
    section("Base XP distribution", sorted((f"XP {k}", v) for k, v in c.items()))
    set_widths(ws4, [4, 60, 14])

    # ---------------- Data Dictionary ----------------
    ws5 = wb.create_sheet("Data Dictionary")
    ws5.append(["Column", "Meaning"])
    style_header(ws5, 2)
    dd = [
        ("NOTE — catalog vs runtime", "This audit documents the full Aurisar catalog (both layers). A few catalog fields are not carried onto the in-app exercise object by the current loader: Default Sets/Reps/Duration and Secondary PB for every exercise, and XP Input Formula, Superset/Interval Eligible, Resistance Level, Primary PB Metric, and Mark As PB for database-only exercises. Those cells show the catalog data even though the app drops them at runtime."),
        ("Dup Group", "Exercises sharing a Dup Group ID are high-confidence duplicates of each other (same name after normalization, a parenthetical alias like 'Bodyweight Squat (air squat)' vs 'Air Squat', or an obvious spelling/format variant). Grouped rows are sorted adjacent and color-banded; all rows are kept."),
        ("Suggested Canonical", "Within a duplicate group, the member suggested to keep (prefers non-archived entries, presence in both catalog layers, and the most complete metadata)."),
        ("ID", "Unique exercise identifier (slug) used throughout the app."),
        ("Name", "Display name."),
        ("Icon", "Emoji shown in the app for this exercise."),
        ("Category", "Primary scoring category: strength / cardio / flexibility / endurance. Drives the class-bonus fallback in XP math."),
        ("Secondary Category", "Optional second category (only a handful of exercises use it)."),
        ("Exercise Type", "Comma-separated style tags (e.g. 'calisthenics, strength', 'yoga, stretching'). Flagged when none of the tags corresponds to the scoring Category."),
        ("Muscle Group", "Primary muscle group key (abs, back, bicep, calves, cardio, chest, forearm, full_body, glutes, legs, shoulder, tricep)."),
        ("Muscles (display)", "Human-readable muscle label shown in the UI."),
        ("Equipment", "Required equipment (bodyweight, dumbbell, barbell, machine, cable, kettlebell, band, etc.)."),
        ("Difficulty", "Beginner / Intermediate / Advanced."),
        ("Description", "Instructional description shown in the exercise detail view."),
        ("Base XP", "Base XP awarded before multipliers (XP = BaseXP x class % x volume x distance/pace/weight/heart-rate bonuses, per calcExXP in src/utils/xp.js)."),
        ("Class Affinity", "The single class the exercise is themed to."),
        ("Best XP Class(es)", "Class(es) with the highest XP % for this exercise."),
        ("<Class> XP %", "Per-class XP multiplier shown as a percentage (e.g. 112% = that class earns 1.12x Base XP on this exercise). 100% = neutral. Blank = no multiplier defined for that class (the app then falls back to the class's category bonus)."),
        ("XP Input Formula", "Which logged inputs feed the XP volume term (e.g. 'Sets x Reps', 'Sets x Reps x Weight')."),
        ("PB Type / PB Tier / Primary PB Metric / Secondary PB / Mark As PB", "Personal-best tracking: what metric counts as a PB for this exercise."),
        ("WOD Viable", "Eligible for Workout-of-the-Day generation."),
        ("Compound / Isolation", "Multi-joint vs single-joint movement (mutually exclusive; rows with both are flagged)."),
        ("Calisthenics / Olympic / Plyometric", "Movement style flags."),
        ("Tracks Weight / Distance / Incline-Speed / Resistance Level", "Which inputs the logging UI offers for this exercise."),
        ("Superset / Interval Eligible", "Whether the exercise can be used in supersets / interval blocks."),
        ("Default Sets / Reps / Duration", "Pre-filled logging defaults, where defined (catalog database field; the current app loader does not copy these onto runtime exercise objects)."),
        ("Archived", "The catalog database marks this exercise as archived/retired — but the app currently loads archived entries anyway, so these are still live in the app (flagged on the Inconsistencies tab)."),
        ("In App Bundle / In Database", "Which catalog layer(s) the exercise lives in. The app ships a bundled list and patches/extends it from the exercise database at startup; both layers together form the single Aurisar catalog audited here."),
        ("Issues", "All inconsistency findings touching this exercise (details on the Inconsistencies tab)."),
    ]
    for k, v in dd:
        ws5.append([k, v])
        ws5.cell(row=ws5.max_row, column=2).alignment = Alignment(wrap_text=True)
    set_widths(ws5, [44, 130])
    ws5.freeze_panes = "A2"

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"total={n_total} bundle={n_bundle} db={n_db} both={n_both} archived={n_arch}")
    print(f"dup_groups={len(groups)} dup_members={dup_members}")
    print(f"findings={len(findings)} by severity={dict(sev_counts)}")


if __name__ == "__main__":
    main()
