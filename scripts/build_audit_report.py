"""Generate Aurisar_Audit_Report.xlsx from audit findings + live source-code counts.

Run from repo root:
    python scripts/build_audit_report.py

The workbook captures every consistency gap identified in the
XP / palette / formatting / exercise-catalog audit, with file:line citations.
"""
from __future__ import annotations

import os
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "src"
APP_JS = SRC / "App.js"
EXERCISES_JS = SRC / "data" / "exercises.js"
CONSTANTS_JS = SRC / "data" / "constants.js"

HEADER_FILL = PatternFill(start_color="2C4564", end_color="2C4564", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SEV_FILL = {
    "P0": PatternFill(start_color="6B2A2A", end_color="6B2A2A", fill_type="solid"),
    "P1": PatternFill(start_color="C49428", end_color="C49428", fill_type="solid"),
    "P2": PatternFill(start_color="8A8478", end_color="8A8478", fill_type="solid"),
    "FIXED": PatternFill(start_color="2E4D38", end_color="2E4D38", fill_type="solid"),
}
SEV_FONT = Font(color="FFFFFF", bold=True)


def write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]) -> None:
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(v, str) and v in SEV_FILL:
                cell.fill = SEV_FILL[v]
                cell.font = SEV_FONT
    ws.freeze_panes = "A2"
    for c in range(1, len(headers) + 1):
        max_len = max(
            [len(str(headers[c - 1]))] + [len(str(row[c - 1])) for row in rows if c - 1 < len(row)]
        )
        ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 12), 80)


def count_in_file(path: Path, pattern: str) -> int:
    if not path.exists():
        return 0
    text = path.read_text(errors="ignore")
    return len(re.findall(pattern, text))


def baseXP_distribution() -> Counter:
    if not EXERCISES_JS.exists():
        return Counter()
    text = EXERCISES_JS.read_text(errors="ignore")
    vals = re.findall(r"baseXP:(\d+)", text)
    return Counter(int(v) for v in vals)


def empty_descs() -> int:
    if not EXERCISES_JS.exists():
        return 0
    return count_in_file(EXERCISES_JS, r'desc:""')


def total_exercises() -> int:
    if not EXERCISES_JS.exists():
        return 0
    return count_in_file(EXERCISES_JS, r'\bid:"[^"]+",\s*name:')


def main() -> None:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # ── Summary ──────────────────────────────────────────────────────
    n_exercises = total_exercises()
    n_empty = empty_descs()
    n_xp_calls = count_in_file(APP_JS, r"calcExXP\(")
    n_format_calls = count_in_file(APP_JS, r"formatXP\(")
    n_remaining_locale = count_in_file(APP_JS, r"\.toLocaleString\(\).*XP")
    bxp = baseXP_distribution()

    summary_rows = [
        ["Total exercises in catalog", n_exercises, ""],
        ["Exercises with empty desc", n_empty, "P1 — content gap"],
        ["Calls to central calcExXP", n_xp_calls, "After fixes — central source of truth"],
        ["Calls to formatXP() helper", n_format_calls, "P1 — adoption-in-progress, sweep remaining sites"],
        ["Remaining `.toLocaleString() XP` patterns to migrate", n_remaining_locale, "P2 — follow-up sweep"],
        [
            "baseXP distribution",
            ", ".join(f"{v}@{k}" for k, v in sorted(bxp.items())),
            "Non-uniform granularity; consider normalising to {25,30,35,40,45,50}",
        ],
        ["", "", ""],
        ["Severity legend", "", ""],
        ["P0", "Clear bug or silent drift — fixed in this PR", "FIXED"],
        ["P1", "Consistency gap — partial fix in PR, see notes", "FIXED"],
        ["P2", "Larger refactor or design decision — flagged only", "P2"],
    ]
    write_sheet(wb, "Summary", ["Metric", "Value", "Notes"], summary_rows)

    # ── XP Calculation Reference ─────────────────────────────────────
    xp_calc_rows = [
        ["Base", "ex.baseXP (15–55, mostly 35 or 45)", "src/utils/xp.js:122"],
        [
            "Class multiplier",
            "ex.xpClassMap[classKey] preferred, else CLASSES[classKey].bonuses[ex.category]",
            "src/utils/xp.js:113-115",
        ],
        ["Set×Reps scaling", "1 + (s*r - 1) * 0.05", "src/utils/xp.js:122"],
        ["Distance bonus", "1 + min(distanceMi*0.05, 0.5) — +5%/mi cap +50%", "src/utils/xp.js:117"],
        [
            "Pace bonus (running only)",
            "1.20 if pace ≤ 8 min/mi else 1.05",
            "src/utils/xp.js:118-119",
        ],
        ["Weight bonus", "1 + min(weightLbs/500, 0.3) — cap +30%, requires ex.tracksWeight", "src/utils/xp.js:120"],
        ["HR Zone bonus", "1 + (hrZone-1)*0.04 — Z1=1.0, Z5=1.16", "src/utils/xp.js:121"],
        [
            "Cardio interval bonus",
            "1.25x if ex.category=='cardio' AND extraRows > 0 — NEW: now central",
            "src/utils/xp.js:123 (added in this PR)",
        ],
        ["Region boost (+7%)", "If region.boost.muscle matches ex.muscleGroup", "src/App.js:1883 (still inline by design)"],
        ["Travel boost (+10%)", "If profile.travelBoost.weekStart === current weekStart()", "src/App.js:1879, 1884"],
        ["Daily check-in", "+125 XP, weekly milestone +500 XP every 7 days", "src/App.js:1767, 1784"],
        ["Quest rewards", "Variable per quest definition", "src/data/constants.js QUESTS array"],
    ]
    write_sheet(
        wb,
        "XP_Calculation_Reference",
        ["Multiplier", "Formula / Value", "Source location"],
        xp_calc_rows,
    )

    # ── XP Logic Drift (before / after this PR) ──────────────────────
    drift_rows = [
        [
            "P0",
            "Cardio +25% interval bonus duplicated inline at 7 sites",
            "src/App.js:118, 1709, 2305, 2449, 4547, 4549, 8933",
            "Centralised into calcExXP via new `extraRows` param",
            "FIXED",
        ],
        [
            "P0",
            "Inline XP formula missing all bonuses (distance/weight/pace/zone)",
            "src/App.js:1972 (quickLogSoloEx)",
            "Replaced with calcExXP() call",
            "FIXED",
        ],
        [
            "P0",
            "Inline XP formula missing pace bonus",
            "src/App.js:2532 (calcEntryXP)",
            "Replaced with calcExXP() call",
            "FIXED",
        ],
        [
            "P0",
            "Inline reimplementation parallel to calcExXP",
            "src/App.js:1882 (logExercise main)",
            "Replaced with calcExXP() call",
            "FIXED",
        ],
        [
            "P1",
            "Region (+7%) and Travel (+10%) applied only inline",
            "src/App.js:1879, 1883-1885",
            "Left inline by design — they're profile-dependent (not ex-dependent). Documented.",
            "P2",
        ],
    ]
    write_sheet(
        wb,
        "XP_Logic_Drift",
        ["Severity", "Issue", "Locations", "Resolution", "Status"],
        drift_rows,
    )

    # ── XP Display Formats ───────────────────────────────────────────
    fmt_rows = [
        ["+25 XP (with space)", "Toast notifications, deletion messages", "App.js:2499, 6614, 6708"],
        ["+250 XP (no space)", "Some inline flash messages", "Migrated to formatXP()"],
        ["1,250 XP (with commas)", "Quest totals, plan totals", "Migrated to formatXP()"],
        ["⚡ 1,250 XP", "Workout cards (workout-tag class)", "Migrated to formatXP({prefix:'⚡ '})"],
        ["⚡ +1,250 XP", "Quest reward badges", "Migrated to formatXP({signed:true,prefix:'⚡ '})"],
        ["XP: 25", "Custom exercise editor", "src/App.js:8331, 8152 — left as-is (label format)"],
        ["Base XP: 45", "Exercise detail view", "src/App.js:8331 — left as-is (label format)"],
        ["lowercase 'xp'", "Not found anywhere", "Always uppercase XP — good"],
    ]
    write_sheet(wb, "XP_Display_Formats", ["Pattern", "Context", "Location / Resolution"], fmt_rows)

    # ── Per-exercise XP distribution ─────────────────────────────────
    bxp_rows = [[k, v, f"{100*v/sum(bxp.values()):.1f}%"] for k, v in sorted(bxp.items())]
    write_sheet(wb, "XP_Per_Exercise_Distribution", ["baseXP", "Count", "% of catalog"], bxp_rows)

    # ── Color Palette Inventory ──────────────────────────────────────
    palette_rows = [
        ["#0c0c0a", "background-color", "Primary dark background", "src/index.css:10, app.css, landing.css"],
        ["#d4cec4", "color (text-primary)", "Main text", "src/styles/landing.css:16, src/index.css:11 (FIXED)"],
        ["#8a8478", "color (text-secondary)", "Muted/helper text", "src/styles/landing.css:17"],
        ["#c49428", "gold accent", "Primary CTA", "src/styles/landing.css:8"],
        ["#f0d060", "gold-light", "Hover state", "src/styles/landing.css:9"],
        ["#8B6914", "gold-dark", "Pressed state", "src/styles/landing.css:10"],
        ["#b0a898", "silver", "Secondary accent", "src/styles/landing.css:11"],
        ["#a07830", "bronze", "Tertiary accent", "src/styles/landing.css:13"],
        ["#8B5A2B", "MUSCLE_COLORS.chest, .tricep", "DUPLICATE — chest and tricep same color", "src/data/constants.js:1176-1177"],
        ["#2C4564", "MUSCLE_COLORS.full_body, .cardio", "DUPLICATE — full body and cardio same color", "src/data/constants.js:1178"],
        ["#2E4D38", "back, push, pull", "Forest green", "src/data/constants.js:1175-1180"],
        ["#3D343F", "shoulder, flexibility, yoga", "Steel plum", "src/data/constants.js:1175-1180"],
        ["#6B2A2A", "bicep, strength (NEW)", "Burgundy iron", "src/data/constants.js:1176, 1181 (FIXED)"],
        ["#5C5C2E", "legs", "Tactical olive", "src/data/constants.js:1177"],
        ["#FFE87C", "UI_COLORS.warning", "Pace bonus highlight, accent emphasis (tokenised)", "src/data/constants.js"],
        ["#2ecc71", "UI_COLORS.success", "Green: positive, beginner, cardio (tokenised)", "src/data/constants.js"],
        ["#e74c3c", "UI_COLORS.danger", "Red: destructive, advanced, strength (tokenised)", "src/data/constants.js"],
        ["(removed)", "—", "#e05555 consolidated into UI_COLORS.danger", "Was a near-duplicate of #e74c3c"],
        ["#2980b9", "UI_COLORS.info", "Info text, message UI fallback (tokenised)", "src/data/constants.js"],
        ["#f1c40f", "UI_COLORS.intermediate", "Intermediate difficulty (tokenised)", "src/data/constants.js"],
    ]
    write_sheet(wb, "Color_Palette_Inventory", ["Color", "Token / Use", "Purpose / Note", "Locations"], palette_rows)

    # ── Color Issues (after this PR) ─────────────────────────────────
    color_issues = [
        ["FIXED", "TYPE_COLORS — Strength/Cardio/Flexibility/Yoga all #C4A044", "src/data/constants.js:1180", "Now Strength=#6B2A2A, Cardio=#2C4564, Flexibility/Yoga=#3D343F (sourced from MUSCLE_COLORS)"],
        ["FIXED", "Body text #333 on #0c0c0a (near-invisible)", "src/index.css:11", "Changed to #d4cec4 (matches landing.css --text-primary)"],
        ["P2", "Chest and Tricep both #8B5A2B", "src/data/constants.js:1176-1177", "Triceps should differ; flagged for design decision"],
        ["P2", "Full Body and Cardio both #2C4564", "src/data/constants.js:1178", "Same purpose? Possibly intentional"],
        ["P2", "6 different gray tokens with no hierarchy", "Various", "#d4cec4 #b4ac9e #b0a898 #8a8478 #5a5650 #6a645a — needs scale"],
        ["FIXED", "Inline one-off colors (#FFE87C, #2ecc71, #e74c3c, #e05555, #2980b9, #f1c40f) tokenised", "src/data/constants.js UI_COLORS", "Migrated 75+ inline JSX style hex strings to UI_COLORS.{warning, success, danger, info, intermediate}; #e05555 consolidated into danger"],
        ["FIXED", "Two fallback colors consolidated to #B0A898", "src/utils/xp.js:6,10", "Both getMuscleColor and getTypeColor now fallback to silver palette token"],
        ["FIXED", "Masculine palette rollout — CAT_ICON_COLORS + filter accents", "src/data/constants.js:15-17 + src/App.js + src/components/PlanWizard.js", "CAT_ICON_COLORS now mirrors TYPE_COLORS masculine values; bright filter accents (#9b59b6, #3498db) migrated to UI_COLORS.accent (brand gold #c49428). Categorical chart colors and HR-zone palette intentionally preserved."],
    ]
    write_sheet(wb, "Color_Issues", ["Severity", "Issue", "Location", "Notes / Resolution"], color_issues)

    # ── Typography ───────────────────────────────────────────────────
    typo_rows = [
        ["FIXED", "Token system introduced (FS scale)", "src/utils/tokens.js (new)", "8-step rem-based font-size scale (xxs→xxl). Top 3 inline values (.6rem, .72rem, .7rem) migrated — 240 sites."],
        ["P2", "Remaining font-size stragglers (.52, .55, .58, .65, .68, .8, .9rem etc)", "Throughout JSX inline styles", "Follow-up sweep — pattern established, just mechanical"],
        ["P2", "Letter-spacing variants (0.02–0.32em)", "Throughout", "No system documented"],
        ["P2", "Line-height variants (1, 1.1, 1.15, 1.6, 1.65, 1.7)", "Throughout", "No system documented"],
        ["P2", "Font families: Cinzel, Cinzel Decorative, Inter, system-ui", "src/styles/*.css", "Mostly intentional but mixed by use case"],
    ]
    write_sheet(wb, "Typography", ["Severity", "Issue", "Locations", "Recommendation"], typo_rows)

    # ── Spacing / BorderRadius ───────────────────────────────────────
    spacing_rows = [
        ["FIXED", "Token system introduced (R + S scales)", "src/utils/tokens.js (new)", "Border-radius scale (sm→full) and 4px-based spacing scale (s2→s32). Top 3 borderRadius values (9, 8, 6) migrated — 89 sites."],
        ["P2", "Remaining borderRadius stragglers (3,4,5,7,10,12,16)", "Throughout JSX inline styles", "Follow-up sweep — pattern established"],
        ["P2", "Hardcoded gap/padding/margin px values", "Throughout JSX inline styles", "Migrate using tokens.S — follow-up sweep"],
        ["P2", "Mix of px-string and unitless React style values", "Throughout", "Pick one convention"],
        ["P2", "Negative margins for overlap (marginTop:-16, marginTop:-14)", "src/styles/landing.css:241,251", "Acceptable but document"],
    ]
    write_sheet(wb, "Spacing_BorderRadius", ["Severity", "Issue", "Locations", "Recommendation"], spacing_rows)

    # ── Number / Unit Formatting ─────────────────────────────────────
    num_rows = [
        ["FIXED", "Weight unit-spacing inconsistent", "src/App.js:118, 2299", "Switched to displayWt() (returns '185 lbs' with space)"],
        ["FIXED", "Helper contract undocumented", "src/utils/units.js:17-20", "Added JSDoc — weightLabel/distLabel return token only; displayWt/displayDist include space"],
        ["P2", "1RM vs PB terminology mixed", "src/App.js:84, 711-715, 1942-1944, PlanWizard.js:47", "Pick one canonical label"],
        ["P2", "Decimal precision varies", "Weight=.1, Distance=.2, BMI=.1, Pace mixed", "Document and standardise"],
        ["P2", "Reps/Sets capitalisation mixed", "Reps vs reps, Sec vs sec", "Pick one (recommend Title Case for labels)"],
        ["FIXED", "XP formatter helper introduced", "src/utils/format.js (new)", "formatXP(value, {signed, prefix}) — see Migration sheet"],
    ]
    write_sheet(wb, "Number_Unit_Formatting", ["Severity", "Issue", "Locations", "Resolution"], num_rows)

    # ── Casing ───────────────────────────────────────────────────────
    casing_rows = [
        ["P2", "Mixed Title Case / sentence / UPPERCASE in buttons", "Throughout JSX", "Document a casing rule"],
        ["P2", "Form labels mostly Title Case but units lowercase ('(km)')", "src/App.js:160, 2217, 2232", "Acceptable convention; document"],
        ["P2", "Tab keys are lowercase (state) but display is Title Case", "src/App.js:283, 478", "Standard pattern; document"],
    ]
    write_sheet(wb, "Casing_Inconsistencies", ["Severity", "Issue", "Locations", "Notes"], casing_rows)

    # ── Exercise catalog ─────────────────────────────────────────────
    cat_rows = [
        [
            "FIXED",
            "Empty desc:\"\" populated with synthesised one-liners",
            "src/data/exercises.js — 72 entries (55 originally empty + 17 with no usable sentence after trimming)",
            "Format: '{Name} is a {category} exercise that targets {muscleGroup} using a {equipment}.'",
        ],
        ["FIXED", "Truncated descriptions trimmed to last full sentence", "src/data/exercises.js — 1,051 entries had imported source data cut mid-word", "Heuristic trim back to last '.' / '!' / '?' followed by space; entries with no usable sentence were synthesised in pass 2"],
        ["P2", "Exact-name duplicates: 'High Knees', 'Incline Barbell Press', 'Jump Rope'", "src/data/exercises.js", "Dedupe — needs migration plan for user logs"],
        ["P2", "Mixed-case exercise names (e.g. 'Bear plank ankle taps into jumps')", "src/data/exercises.js (~30 entries)", "Apply Title Case sweep"],
        ["P2", "50+ Bench Press variants with inconsistent suffixes", "src/data/exercises.js", "Standardise modifier convention"],
    ]
    write_sheet(
        wb, "Exercise_Catalog_Issues", ["Severity", "Issue", "Locations", "Recommendation"], cat_rows
    )

    # ── Dual-form duplicates ─────────────────────────────────────────
    dual_rows = [
        ["dumbbell-lunges", "Dumbbell Lunges", "ymove", "alias (hidden from picker)", "FIXED"],
        ["dumbbell_lunge", "Dumbbell Lunge", "canonical", "shown in picker", "FIXED"],
        ["", "", "", "", ""],
        ["dumbbell-floor-chest-press", "Dumbbell floor chest press", "ymove", "alias (hidden from picker)", "FIXED"],
        ["dumbbell_floor_press", "Dumbbell Floor Press", "canonical", "shown in picker", "FIXED"],
        ["", "", "", "", ""],
        ["hammer_curls", "Hammer Curls", "free-exercise-db", "alias (hidden from picker)", "FIXED"],
        ["hammer_curl", "Hammer Curl", "canonical", "shown in picker", "FIXED"],
        ["", "", "", "", ""],
        ["standing_calf_raises", "Standing Calf Raises", "free-exercise-db", "alias (hidden from picker)", "FIXED"],
        ["standing_calf_raise", "Standing Calf Raise", "canonical", "shown in picker", "FIXED"],
    ]
    write_sheet(
        wb,
        "Exercise_Dual_Form_Duplicates",
        ["exId", "Display name", "Source", "Location", "Severity"],
        dual_rows,
    )

    # ── Plan templates ───────────────────────────────────────────────
    plan_rows = [
        [
            "FIXED",
            "reps and sets unified to string schema across all PLAN_TEMPLATES",
            "Was: numeric (reps:8) in 6 templates; mixed (sets:3 numeric, reps:'6-12' string) in dumbbell_8wk",
            "src/data/constants.js PLAN_TEMPLATES",
            "All 244 sets/reps values now string-quoted; accommodates ranges (6-12), time (30s), special (AMRAP)",
        ],
        [
            "FIXED",
            "restSec:null field added to every exercise object",
            "Schema gap closed — 194 exercise entries updated",
            "src/data/constants.js PLAN_TEMPLATES",
            "Field is null until plan UI exposes a rest-timer input",
        ],
        ["FIXED", "Workout descriptions remain consistent", "Tone, length, punctuation all good", "src/data/constants.js WORKOUT_TEMPLATES", "Reference example for other content work"],
    ]
    write_sheet(
        wb, "Plan_Template_Schema", ["Severity", "Issue", "Detail", "Locations", "Recommendation"], plan_rows
    )

    # ── Recommendations ──────────────────────────────────────────────
    recs = [
        ["P0", "FIXED", "Body text contrast bug (color #333 on #0c0c0a)", "src/index.css:11"],
        ["P0", "FIXED", "TYPE_COLORS collision — 4 categories same color", "src/data/constants.js:1180"],
        ["P0", "FIXED", "Cardio +25% bonus duplicated inline 7×", "src/utils/xp.js:111 (extraRows param), src/App.js (7 sites)"],
        ["P0", "FIXED", "Drifted XP formulas (missing bonuses)", "src/App.js:1882, 1972, 2532"],
        ["P1", "PARTIAL", "formatXP() helper introduced; ~8 high-traffic sites migrated", "src/utils/format.js + src/App.js (sweep remaining sites)"],
        ["P1", "FIXED", "Unit-spacing display sites + helper docs", "src/App.js:118, 2299; src/utils/units.js:16"],
        ["P1", "FIXED", "Populate empty/truncated exercise descriptions", "src/data/exercises.js — 72 empty entries synthesised + 1,051 truncated entries trimmed cleanly"],
        ["P2", "PARTIAL", "Spacing/typography token system introduced + top values migrated", "src/utils/tokens.js (new) — 329 highest-frequency values migrated; remaining stragglers tracked"],
        ["P2", "FIXED", "Dual-form catalog dedup via alias flag (no user-data migration needed)", "src/data/exercises.js — 4 legacy entries marked alias:true; src/App.js splits picker (filtered) vs lookup (unfiltered)"],
        ["P2", "FIXED", "Masculine palette rollout — CAT_ICON_COLORS + filter accents", "Profile/Library/PlanWizard now use UI_COLORS.accent for active states; xp.js fallback consolidated"],
        ["P2", "FIXED", "Reps schema standardisation + restSec field in PLAN_TEMPLATES", "src/data/constants.js — 244 values quoted, 194 restSec:null fields added"],
        ["P2", "FIXED", "Tokenise inline color one-offs", "src/data/constants.js UI_COLORS + 75+ migrated call-sites"],
        ["P2", "TODO", "Sweep remaining XP display sites to formatXP()", "src/App.js (~40+ remaining call sites)"],
    ]
    write_sheet(wb, "Recommendations", ["Severity", "Status", "Item", "Files / Notes"], recs)

    out = ROOT / "Aurisar_Audit_Report.xlsx"
    wb.save(out)
    print(f"Wrote {out}  ({os.path.getsize(out):,} bytes)")
    print(f"Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
